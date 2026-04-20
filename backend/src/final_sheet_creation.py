import os
import json
import math
import numpy as np
import pandas as pd
import string
import warnings
from datetime import datetime
from openpyxl.styles import Alignment, Border, Side, Font

warnings.filterwarnings("ignore")

def uuid_sequence_for_additional(business_matched_records, uuid_for_additional):
    additional_uuid = []
    max_row_count_len = len(str(business_matched_records.shape[0]))
    for i in range(1, business_matched_records.shape[0] + 1):
        x = uuid_for_additional + '0' * (max_row_count_len - len(str(i))) + str(i)
        additional_uuid.append(x)
    return additional_uuid

def single_column_conversion(dataset, columns_name):
    output_string = []
    dataset.fillna('', inplace=True)
    valid_columns = [col for col in columns_name if col in dataset.columns]
    
    if not valid_columns:
        return [''] * len(dataset)
        
    for mylist in dataset[valid_columns].values:
        string_output = ''
        for item in mylist:
            val = str(item).strip()
            if val.lower() not in ('', '-', 'nan', 'none'):
                string_output = val
                break
        output_string.append(string_output)
        
    return output_string

def data_reconstruction(dataset, columns_name, name):
    main_contact, index_values = [], []
    combine_details = single_column_conversion(dataset=dataset, columns_name=columns_name)  
    
    for ref_value, ref_index in zip(combine_details, dataset.index):
        sample_dataset = dataset[columns_name]
        common_list_details = []
        
        for remaining_details in sample_dataset.iloc[ref_index, :].values:
            if len(str(remaining_details)) >= 2:          
                if ref_value != remaining_details:
                    common_list_details.append(remaining_details)
        main_contact.append((common_list_details, ref_index)) 
        
    values_data, index_values = zip(*main_contact)
    updated_length = []
    for max_column_count in values_data:
        if len(max_column_count) not in updated_length:
            updated_length.append(len(max_column_count))  

    mydata = pd.DataFrame(values_data, columns=[f'{name}_{p}' for p in range(2, max(updated_length) + 2) if max(updated_length) > 0])

    if mydata.shape[1] <= 1:
        final_output = pd.DataFrame(combine_details, columns=[f'{name}'])
    else:
        final_output = pd.DataFrame(combine_details, columns=[f'{name}_{1}'])
         
    concated_output = pd.concat([final_output, mydata], axis=1)
    concated_output.fillna('', inplace=True)
    concated_output = concated_output.mask(concated_output == '')
    concated_output[name] = concated_output.apply(lambda x: ';'.join(x.dropna().astype(str).values), axis=1)
    df = pd.DataFrame(concated_output[name])
    df[f'{name}_index'] = index_values
    return df

dropdown_datatype_columns = {
    'Primary SIC Code and Sector': 'Number',
    'PRIMARY SIC SECTOR': 'Text(Small)',
    'Primary NAICS Code and Sector': 'Number',
    'PRIMARY NAICS SECTOR': 'Text(Small)'
}

def isfloat(num):
    try:
        float(num)
        return True
    except ValueError:
        return False
    
def largest_num_list(city_df, updated_col_name, index):
    max_number_value_list = []
    for data in city_df[updated_col_name[index]]:
        data = str(data).replace(';', ' ')
        if data != '':
            data1 = data.split(" ")
            max_number_value = []
            for i in data1:
                if i.isnumeric():
                    max_number_value.append(int(i))
                elif isfloat(i):
                    max_number_value.append(math.ceil(float(i)))
                elif not i.isalpha():
                    res = ''.join(filter(lambda x: x.isdigit(), i))
                    if res:
                        max_number_value.append(int(res))
            if len(max_number_value) >= 1:
                max_number_value_list.append(max(max_number_value))
            else:
                max_number_value_list.append('')
        else:
            max_number_value_list.append('')
    return max_number_value_list

def earliest_year_list(city_df, updated_col_name, index):
    min_year_list = []
    for data in city_df[updated_col_name[index]]:
        data = str(data).replace(',', ' ').replace('-', ' ').replace('/', ' ').replace(':', ' ').replace(';', ' ')
        if data != '':
            data1 = data.split(" ")
            year = []
            for i in data1:
                if i.isnumeric():
                    if 9999 >= int(i) >= 1000:
                        year.append(i)
            if len(year) >= 1:
                min_year_list.append(min(year))
            else:
                min_year_list.append('')
        else:
            min_year_list.append('')
    return min_year_list

def earliest_date_list(city_df, updated_col_name, index):
    earliest_date_list = []
    for data in city_df[updated_col_name[index]]:
        data = str(data).replace(';', ' ')
        if data != '':
            data1 = data.split(" ")
            date1 = []
            for i in data1:
                try:
                    date1.append(datetime.strptime(i, '%m/%d/%Y'))
                except ValueError:
                    pass
            if len(date1) == 1:
                earliest_date_list.append(date1[0].strftime("%m/%d/%Y"))
            elif len(date1) > 1:
                x = date1[0]
                for j in date1[1:]:
                    if x > j:
                        x = j
                earliest_date_list.append(x.strftime("%m/%d/%Y"))
            else:
                earliest_date_list.append('')
        else:
            earliest_date_list.append('') 
    return earliest_date_list

def datatype_by_column_data(dataset, new_field_col):
    if f'{new_field_col}_1' in dataset.columns:
        column_data = dataset[f'{new_field_col}_1']
    elif new_field_col in dataset.columns:
        column_data = dataset[new_field_col]
    else:
        return 'Text(Small)'

    if pd.to_numeric(column_data, errors='coerce').notnull().all():
        return 'Number'
    elif column_data.astype(str).str.match(r'^\d{4}$').all():
        return 'Year'
    elif pd.to_datetime(column_data, errors='coerce').notnull().all():
        return 'Date'
    else:
        value_lengths = column_data.astype(str).str.len()
        max_length = value_lengths.max()
        if pd.isna(max_length):
            return 'Text(Small)'
        elif max_length < 256:
            return 'Text(Small)'
        elif max_length < 4096:
            return 'Text(Medium)'
        else:
            return 'Text(Large)'

def get_custom_matched_records(dataset, city_field_mapping, bludot_field_mapping, method, uuid_for_additional, new_fields, raw_sheet, city_records, filename_output):
    # =========================================================================
    # ID STANDARDIZER SAFEGUARD
    # =========================================================================
    if 'ID' not in dataset.columns:
        if 'UUID' in dataset.columns: dataset['ID'] = dataset['UUID']
        elif 'Id' in dataset.columns: dataset['ID'] = dataset['Id']
        elif 'id' in dataset.columns: dataset['ID'] = dataset['id']
        else: dataset['ID'] = ''
    # =========================================================================

    if method == True:
        dataset['ID'] = uuid_sequence_for_additional(dataset, uuid_for_additional)
        if len(bludot_field_mapping) != 0:
            for column in bludot_field_mapping:
                dataset[column] = ''
    
    datatype = []
    new_field_datatype = []
    updated_col_name = []
    custom_column_name = []
    default_or_user_value = []
    city_df = pd.DataFrame()
    custom_output_sheet = pd.DataFrame()
    
    for col in city_field_mapping:
        col_duplicates = [c for c in dataset.columns if c.startswith(f'{col}_')]
        df = dataset.loc[:, col_duplicates]
        df.fillna('', inplace=True)
        df = df.mask(df == '')
        city_df[col] = df.apply(lambda x: ';'.join(set(x.dropna().astype(str).values)), axis=1)

    if len(bludot_field_mapping) != 0:
        # COMPLETELY BYPASSING RAW FILE READS - PURELY DATA DRIVEN NOW
        for list_index, custom_col_name in enumerate(bludot_field_mapping):
            
            # Infer the datatype directly from the column values
            col_datatype = dropdown_datatype_columns.get(
                custom_col_name, 
                datatype_by_column_data(dataset, city_field_mapping[list_index])
            )
            
            # Standardize all as User Defined
            updated_col_name.append(city_field_mapping[list_index])
            custom_column_name.append(custom_col_name)
            datatype.append(col_datatype)
            default_or_user_value.append('User Defined')
    
    if len(new_fields):
        for new_field_col in new_fields:
            new_field_datatype.append(datatype_by_column_data(dataset, new_field_col))
            
    for index, col in enumerate(default_or_user_value):
        if col == 'Default':
            custom_output_sheet[updated_col_name[index]] = dataset[updated_col_name[index]]
        else:
            if datatype[index] == 'Number':
                custom_output_sheet[updated_col_name[index]] = largest_num_list(city_df, updated_col_name, index)
            elif datatype[index] == 'Year':
                custom_output_sheet[updated_col_name[index]] = earliest_year_list(city_df, updated_col_name, index)
            elif datatype[index] == 'Date':
                custom_output_sheet[updated_col_name[index]] = earliest_date_list(city_df, updated_col_name, index)
            else:
                custom_output_sheet[updated_col_name[index]] = city_df[updated_col_name[index]]
                
            if updated_col_name[index] in new_fields:
                new_col_name = updated_col_name[index] + datatype[index]
                custom_output_sheet.rename(columns={updated_col_name[index]: new_col_name}, inplace=True)
                updated_col_name[index] = new_col_name

    new_fields_df = pd.DataFrame()
    for col in new_fields:
        col_duplicates = [c for c in dataset.columns if c.startswith(f'{col}_')]
        new_fields_col = dataset.loc[:, col_duplicates]
        new_fields_col.fillna('', inplace=True)
        new_fields_col = new_fields_col.mask(new_fields_col == '')
        new_fields_df[col] = new_fields_col.apply(lambda x: ';'.join(set(x.dropna().astype(str).values)), axis=1)
        
    for index, col_datatype in enumerate(new_field_datatype):
        if col_datatype == 'Number':
            custom_output_sheet[new_fields[index]] = largest_num_list(new_fields_df, new_fields, index)
        elif col_datatype == 'Year':
            custom_output_sheet[new_fields[index]] = earliest_year_list(new_fields_df, new_fields, index)
        elif col_datatype == 'Date':
            custom_output_sheet[new_fields[index]] = earliest_date_list(new_fields_df, new_fields, index)
        else:
            custom_output_sheet[new_fields[index]] = new_fields_df[new_fields[index]]

        updated_col_name.append(new_fields[index])
        custom_column_name.append(new_fields[index])
        datatype.append(col_datatype)
        default_or_user_value.append('User Defined')
        
    custom_output_sheet.columns = custom_column_name
    custom_output_sheet.loc[-1] = default_or_user_value
    custom_output_sheet = custom_output_sheet.sort_index().reset_index(drop=True)
    
    custom_output_sheet.loc[-1] = datatype
    custom_output_sheet = custom_output_sheet.sort_index().reset_index(drop=True)
    
    custom_output_sheet.insert(0, 'Custom Data Name', ['Custom Data Type', 'ID'] + dataset['ID'].tolist())
    default_values = [index for index, i in enumerate(default_or_user_value) if i == 'Default']
    
    return (custom_output_sheet, default_values)

def format_custom_subsheet(workbook, subsheet_name, default_values):
    worksheet = workbook[subsheet_name]

    for default_value_index in default_values:
        col_start = default_value_index + 1
        col_end = default_value_index + 2
        worksheet.merge_cells(start_row=1, start_column=col_start + 1, end_row=1, end_column=col_end + 1)
        worksheet.cell(row=1, column=col_start + 1).alignment = Alignment(horizontal='center', vertical='center')
        worksheet.merge_cells(start_row=2, start_column=col_start + 1, end_row=2, end_column=col_end + 1)
        worksheet.cell(row=2, column=col_start + 1).alignment = Alignment(horizontal='center', vertical='center')

    for row_num in range(1, 4):
        row = worksheet[row_num]
        for cell in row:
            cell.font = Font(bold=True)
            cell.border = Border(top=Side(border_style='thin'), right=Side(border_style='thin'),
                                 bottom=Side(border_style='thin'), left=Side(border_style='thin'))

def get_country_state(file_path, state_to_country):
    current_dir = os.path.dirname(os.path.abspath(__file__))
    safe_file_path = os.path.join(current_dir, 'country_state_mapping.json')
    with open(safe_file_path, 'r') as json_file:
        country_state_mapping = json.load(json_file)
    for country, states in country_state_mapping.items():
        for state in states.values():
            state_to_country[state] = country
    return state_to_country

def get_country(state, state_to_country):
    return state_to_country.get(state, '')

def change_sheet_format(dataset, original_record_list, updated_record_list, country_state_mapping):
    state_to_country = get_country_state(country_state_mapping, {})
    column_mapping = {old: new for old, new in zip(original_record_list, updated_record_list)}    
    dataset.rename(columns=column_mapping, inplace=True)
    if 'State' in dataset.columns:
        country_values = dataset['State'].apply(lambda state: get_country(state, state_to_country))
        dataset.insert(dataset.columns.get_loc('State') + 1, 'Country', country_values)
    return dataset

def get_Business_Matched_Records(dataset, city_field_mapping, bludot_field_mapping, original_record_list, updated_record_list, country_state_mapping, city_records, method, uuid_for_additional, source_type='user'):
    """Maps correctly from the actual dataset using smart prioritization and Universal Fallbacks."""
    business_matched_records = pd.DataFrame()
    
    if method == True: 
        for cols_name_add in original_record_list:
            if cols_name_add not in dataset.columns:
                dataset[cols_name_add] = ''
                
    dataset.fillna('', inplace=True)
    
    mapping_field = dict()
    for city_col, bludot_col in zip(city_field_mapping, bludot_field_mapping):
        search_cols = [f"{city_col}_1", city_col, bludot_col, f"{bludot_col}_1"]
        
        # UNIVERSAL FALLBACKS 
        if bludot_col == "Name_x":
            search_cols.extend(["Name", "Business Name", "Account Name"])
        elif bludot_col == "Address1":
            search_cols.extend(["Address1", "Account Address", "Address"])
        elif bludot_col == "City":
            search_cols.extend(["City", "Account City"])
        elif bludot_col == "State":
            search_cols.extend(["State", "Account State"])
        elif bludot_col == "ZipCode":
            search_cols.extend(["ZipCode", "Zipcode", "Account Zipcode", "Zip"])
        elif bludot_col == "PhoneNumber":
            search_cols.extend(["PhoneNumber", "Phonenumber", "Main Account Phone", "Phone"])
            
        mapping_field[bludot_col] = search_cols
        
    for updated_cols_name in original_record_list:
        if updated_cols_name in mapping_field.keys():
            cols_to_search = mapping_field[updated_cols_name]
            updated_single_columns = single_column_conversion(dataset=dataset, columns_name=cols_to_search)
        else:
            if updated_cols_name in dataset.columns:
                updated_single_columns = dataset[updated_cols_name].values
            else:
                updated_single_columns = [''] * len(dataset)

        business_matched_records[updated_cols_name] = updated_single_columns
            
    if method == True:    
       business_matched_records['UUID'] = uuid_sequence_for_additional(business_matched_records, uuid_for_additional)

    business_matched_records = change_sheet_format(dataset=business_matched_records, original_record_list=original_record_list, updated_record_list=updated_record_list, country_state_mapping=country_state_mapping)

    business_matched_records['is_business'] = 'True'    
    business_matched_records['business_source'] = source_type

    # =========================================================================
    # FORCE ESSENTIAL COLUMNS TO EXIST AND PULL FROM BLUDOT DATA
    # =========================================================================
    for req_col in ['DBA Name', 'Business Operational Status']:
        # Find the exact column in the raw merged dataset (ignoring case/spaces)
        actual_col = next((c for c in dataset.columns if str(c).strip().lower() == req_col.lower()), None)
        
        if actual_col:
            # If the Bludot data has it, pull it!
            business_matched_records[req_col] = dataset[actual_col].values
        elif req_col not in business_matched_records.columns:
            # Otherwise, leave it safely blank
            business_matched_records[req_col] = ''

    essential_cols = ['business_source', 'DBA Name', 'Business Operational Status']
    cols = business_matched_records.columns.tolist()
    
    # Pluck them out of their random positions
    for col in essential_cols:
        if col in cols: 
            cols.remove(col)
            
    # Append them neatly at the very end
    for col in essential_cols:
        cols.append(col)
        
    return business_matched_records[cols]

def string_filter_for_abbreviation(strings):
    punctuation_chars = string.punctuation
    strings = strings.lower().translate(str.maketrans(punctuation_chars, ' ' * len(punctuation_chars)))
    return ' '.join(strings.split())

def phone_number_formatting(phone_num):
    return ''.join(char for char in phone_num if char.isdigit())

def get_contact_matched_records(CITY_NAME, city_records, contact_city_columns, column_type, matched_records):
    city_records.fillna('', inplace=True)  
    city_records = city_records.astype(str)  
    updated_contact_city_columns = []  

    for contact_column in contact_city_columns:
        duplicate_column_names = []  
        for column_name in city_records.columns:
            if isinstance(contact_column, list):  
                if column_name.startswith(contact_column[0] + '_'):  
                    duplicate_column_names.append(column_name)
            elif column_name.startswith(contact_column + '_'):  
                duplicate_column_names.append(column_name)
        updated_contact_city_columns.append(duplicate_column_names)  
        
    duplicated_col_max_len = max([len(t) for t in updated_contact_city_columns]) if updated_contact_city_columns else 0
    
    contact_matched_columns_list = []
    contact_matched_columns_type = []
    names_column = []
    city_records.replace('', np.nan, inplace=True)
    
    for j in range(1, duplicated_col_max_len + 1):
        for i in range(len(contact_city_columns)):
            if isinstance(contact_city_columns[i], list):
                name_dup_list = [contact_city_columns[i][0] + '_' + str(j)]
                contact_matched_columns_list.append(name_dup_list[0])
                name_dup_list.append(contact_city_columns[i][1] + '_' + str(j) if contact_city_columns[i][1] != '' else '')
                name_dup_list.append(contact_city_columns[i][2] + '_' + str(j) if contact_city_columns[i][2] != '' else '')
                names_column.append(name_dup_list)
            else:
                contact_matched_columns_list.append(contact_city_columns[i] + '_' + str(j))
            contact_matched_columns_type.append(column_type[i]) 
            
    valid_cols = [c for c in contact_matched_columns_list if c in city_records.columns]
    contact_matched_records = city_records.loc[:, valid_cols]
    contact_matched_records_columns_type = [col_type[0] for col_type in contact_matched_columns_type]
    
    name_email_columns, name_phone_columns, name_index = [], [], []
    for i in range(len(contact_matched_records_columns_type)):
        if contact_matched_records_columns_type[i] == 'Name':
            name_index.append(i)
            for j in range(i + 1, len(contact_matched_records_columns_type)):
                if contact_matched_records_columns_type[j] == 'Email':
                    name_email_columns.append([contact_matched_columns_list[i], contact_matched_columns_list[j]])
                elif contact_matched_records_columns_type[j] == 'Phone':   
                    name_phone_columns.append([contact_matched_columns_list[i], contact_matched_columns_list[j]])
                elif contact_matched_records_columns_type[j] == 'Name':
                    break     
                         
    contact_matched_records.fillna('', inplace=True)
    for x in range(len(name_email_columns) - 1):
        if name_email_columns[x][0] in contact_matched_records.columns and name_email_columns[x][1] in contact_matched_records.columns:
            df = pd.DataFrame(contact_matched_records[name_email_columns[x]]).fillna('')
            for y in range(x + 1, len(name_email_columns)):
                if name_email_columns[y][0] in contact_matched_records.columns and name_email_columns[y][1] in contact_matched_records.columns:
                    df1 = pd.DataFrame(contact_matched_records[name_email_columns[y]]).fillna('').astype(str)
                    for j in range(df1.shape[0]):
                        if str(string_filter_for_abbreviation(df[name_email_columns[x][0]][j])) == str(string_filter_for_abbreviation(df1[name_email_columns[y][0]][j])) or df[name_email_columns[x][0]][j] == '' or df1[name_email_columns[y][0]][j] == '':
                            if str(string_filter_for_abbreviation(df[name_email_columns[x][1]][j])) == str(string_filter_for_abbreviation(df1[name_email_columns[y][1]][j])):
                                contact_matched_records.at[j, name_email_columns[y][1]] = ''

    contact_matched_records.fillna('', inplace=True)        
    for v in [name_email[1] for name_email in name_email_columns]:
        if v in contact_matched_records.columns:
            contact_matched_records.loc[~contact_matched_records[v].str.contains('@'), v] = ''
            contact_matched_records[v] = contact_matched_records[v].replace('', np.nan)                    
    
    for x in range(len(name_phone_columns) - 1):
        if name_phone_columns[x][0] in contact_matched_records.columns and name_phone_columns[x][1] in contact_matched_records.columns:
            df = pd.DataFrame(contact_matched_records[name_phone_columns[x]]).fillna('')                              
            for y in range(x + 1, len(name_phone_columns)):
                if name_phone_columns[y][0] in contact_matched_records.columns and name_phone_columns[y][1] in contact_matched_records.columns:
                    df1 = pd.DataFrame(contact_matched_records[name_phone_columns[y]]).fillna('').astype(str)
                    for j in range(df1.shape[0]):
                        if str(phone_number_formatting(df[name_phone_columns[x][0]][j])) == str(phone_number_formatting(df1[name_phone_columns[y][0]][j])) or str(phone_number_formatting(df[name_phone_columns[x][0]][j])) == '' or str(phone_number_formatting(df1[name_phone_columns[y][0]][j])) == '':
                            if str(phone_number_formatting(df[name_phone_columns[x][1]][j])) == str(phone_number_formatting(df1[name_phone_columns[y][1]][j])):
                                contact_matched_records.at[j, name_phone_columns[y][1]] = ''

    contact_matched_records.fillna('', inplace=True)
    index_num = 0
    if matched_records == 'business_matched' and 'ID' in city_records.columns and 'city_index' in city_records.columns:
        if contact_city_column_index - contact_id_column_index > 1:
            bludot_contact_records = city_records.iloc[:, contact_id_column_index:contact_city_column_index].fillna('')
            updated_bludot_columns_name = [col.split('.')[0] if '.' in col else col for col in bludot_contact_records.columns]
            bludot_contact_records.columns = updated_bludot_columns_name
            
            empty_cols_to_drop = []
            updated_col_with_drop_col = list(bludot_contact_records.columns)
            for col_index, col_name in enumerate(bludot_contact_records.columns):
                bludot_contact_records.iloc[:, col_index] = bludot_contact_records.iloc[:, col_index].replace('', None)
                if np.all(pd.isnull(bludot_contact_records.iloc[:, col_index])):
                    if col_name == 'Name':
                        empty_cols_to_drop.extend([col_index, col_index + 1, col_index + 2])
                    elif col_name not in ['Title', 'Roles']:
                        empty_cols_to_drop.append(col_index)
            
            for index in empty_cols_to_drop:
                if index < len(updated_col_with_drop_col):
                    updated_col_with_drop_col[index] = 'drop_col'
                    
            bludot_contact_records.columns = updated_col_with_drop_col
            bludot_contact_records = bludot_contact_records.drop(columns=[col for col in bludot_contact_records.columns if col == 'drop_col']).fillna('')
            contact_matched_records.fillna('', inplace=True)
            
            bludot_indices = [index for index, col in enumerate(bludot_contact_records.columns) if col == 'Name']
            city_indices = [index for index, col in enumerate(contact_matched_records.columns) if 'Name' in col]
            
            for i in bludot_indices:
                for j in city_indices:
                    for x in range(contact_matched_records.shape[0]):
                        if str(string_filter_for_abbreviation(bludot_contact_records.iloc[x, i])) == str(string_filter_for_abbreviation(contact_matched_records.iloc[x, j])):
                            if bludot_contact_records.iloc[x, i] != "":
                                if j != city_indices[-1]:
                                    if all(y == '' for y in contact_matched_records.iloc[x, j:j+1].tolist()[1:]):
                                        contact_matched_records.iloc[x, j] = ''
                                        
            contact_matched_records = pd.concat([bludot_contact_records, contact_matched_records], axis=1).fillna('')

    return contact_matched_records