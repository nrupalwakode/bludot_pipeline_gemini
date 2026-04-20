"""
Step 5 — Generate Business + Custom + Contact Output Sheets
============================================================
Safely receives city_schema.json and bridges it with final_sheet_creation.py.
Forces 'user' source for additional unmatched city records.
"""

import os
import json
import math
import logging
import pandas as pd
from datetime import datetime
from pathlib import Path
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from ..db.models import City

logger = logging.getLogger(__name__)

def _patch_fsc(fsc):
    def _largest_num_list_fixed(city_df, updated_col_name, index):
        max_number_value_list = []
        for data in city_df[updated_col_name[index]]:
            data = str(data).replace(';', ' ')
            if data != '':
                data1 = data.split(' ')
                max_number_value = []
                for i in data1:
                    if i.isnumeric(): max_number_value.append(int(i))
                    elif fsc.isfloat(i): max_number_value.append(math.ceil(float(i)))
                    elif not i.isalpha():
                        res = ''.join(filter(lambda x: x.isdigit(), i))
                        if res: max_number_value.append(int(res))
                max_number_value_list.append(max(max_number_value) if max_number_value else '')
            else: max_number_value_list.append('')
        return max_number_value_list

    def _earliest_date_list_fixed(city_df, updated_col_name, index):
        from datetime import datetime as _dt
        result = []
        for data in city_df[updated_col_name[index]]:
            data = str(data).replace(';', ' ')
            if data.strip() != '':
                dates = []
                for token in data.split(' '):
                    token = token.strip()
                    if not token: continue
                    try: dates.append(_dt.strptime(token, '%m/%d/%Y'))
                    except (ValueError, TypeError): pass
                result.append(min(dates).strftime('%m/%d/%Y') if dates else '')
            else: result.append('')
        return result

    fsc.largest_num_list  = _largest_num_list_fixed
    fsc.earliest_date_list = _earliest_date_list_fixed

def _load_schema(city_dir: Path) -> dict:
    schema_path = city_dir / "city_schema.json"
    if not schema_path.exists():
        raise FileNotFoundError(f"city_schema.json not found at {schema_path}.")
    with open(str(schema_path)) as f:
        return json.load(f)

def _get_actual_col(df, expected_name):
    if expected_name in df.columns: return expected_name
    if f"{expected_name}_1" in df.columns: return f"{expected_name}_1"
    clean_expected = str(expected_name).strip().lower()
    for col in df.columns:
        clean_col = str(col).strip().lower()
        if clean_col == clean_expected or clean_col == f"{clean_expected}_1": return col
    clean_expected_no_space = clean_expected.replace(' ', '').replace('_', '')
    for col in df.columns:
        clean_col_no_space = str(col).strip().lower().replace(' ', '').replace('_', '')
        if clean_col_no_space == clean_expected_no_space or clean_col_no_space == f"{clean_expected_no_space}1": return col
    return None

def _prepare_data_for_legacy_script(df, schema):
    for col in schema.get("COLUMNS_LIST", []) + schema.get("BUSINESS_MATCHED_BLUDOT_COLUMNS", []):
        if col not in df.columns: df[col] = ""

    all_sources = (
        schema.get("BUSINESS_MATCHED_CITY_COLUMNS", []) +
        schema.get("BUSINESS_CUSTOM_MATCHED_CITY_COLUMNS", []) +
        schema.get("NEW_FIELDS", [])
    )

    for src in all_sources:
        real_col = _get_actual_col(df, src)
        if real_col:
            df[src] = df[real_col]
            df[f"{src}_1"] = df[real_col]
        else:
            df[src] = ""
            df[f"{src}_1"] = "" 
    return df

def run_step5(city: City, db: Session, results_dir: str) -> dict:
    import sys
    backend_dir = str(Path(__file__).resolve().parent.parent)
    if backend_dir not in sys.path: sys.path.append(backend_dir)
        
    try:
        import src.final_sheet_creation as _fsc
        from src.final_sheet_creation import (
            get_Business_Matched_Records, get_custom_matched_records,
            get_contact_matched_records, format_custom_subsheet
        )
    except ImportError as e:
        logger.error(f"Cannot import final_sheet_creation module: {e}")
        raise

    _patch_fsc(_fsc)

    results_path = Path(results_dir)
    city_dir     = results_path.parent
    city_name    = city.name
    schema       = _load_schema(city_dir)

    final_excel  = results_path / "output" / "final_excel"
    final_output = results_path / "output" / "final_output"
    final_excel.mkdir(parents=True, exist_ok=True)
    final_output.mkdir(parents=True, exist_ok=True)

    city_records   = pd.read_excel(str(results_path / "city_data" / "de_duplication_merged.xlsx"), dtype=object)
    bludot_records = pd.read_excel(str(results_path / "bludot_data" / "bludot_concatenated_records.xlsx"), dtype=object)
    raw_sheet      = str(city_dir / "original_record" / city.raw_data_path.split("/")[-1])
    city_records.fillna("", inplace=True)

    final_result = results_path / "output" / "final_result"
    # SMART READ: Don't crash if Step 3 found 0 matches and didn't create the file
    matched_file_path = final_result / f"final_matched_records_for_{city_name}.xlsx"
    if matched_file_path.exists():
        total_match_records = pd.read_excel(str(matched_file_path), dtype=object)
    else:
        logger.warning(f"Matched records file missing for {city_name}. Assuming 0 matches.")
        total_match_records = pd.DataFrame() # Create an empty dataframe so the code survives
    additional_city_match_records = pd.read_excel(str(final_result / f"additional_city_records_for_{city_name}.xlsx"), dtype=object)

    total_match_records.fillna('', inplace=True)
    additional_city_match_records.fillna('', inplace=True)
    total_match_records = total_match_records.astype(str)
    additional_city_match_records = additional_city_match_records.astype(str)

    total_match_records = _prepare_data_for_legacy_script(total_match_records, schema)
    additional_city_match_records = _prepare_data_for_legacy_script(additional_city_match_records, schema)

    now = datetime.now()
    uuid_for_additional = city_name[:3].upper() + str(now.year) + str(now.month) + str(now.day)

    stats = {"business_matched": 0, "business_additional": 0, "custom_matched": 0, "contact_matched": 0}

    # BUSINESS RECORDS
    custom_matched_records = None
    if schema.get("BUSINESS_MATCHED_CITY_COLUMNS"):
        business_matched = get_Business_Matched_Records(
            dataset=total_match_records, city_field_mapping=schema["BUSINESS_MATCHED_CITY_COLUMNS"],
            bludot_field_mapping=schema["BUSINESS_MATCHED_BLUDOT_COLUMNS"], original_record_list=schema["COLUMNS_LIST"],
            updated_record_list=schema["COLUMNS_LIST_UPDATED"], country_state_mapping=schema.get("COUNTRY_STATE_MAPPING", {}),
            method=False, city_records=city_records, uuid_for_additional='', source_type=schema.get("BLUDOT_OR_USER", "bludot"))
            
        # FIXED: Enforces 'user' source type for all additional city records
        add_business_matched = get_Business_Matched_Records(
            dataset=additional_city_match_records, city_field_mapping=schema["BUSINESS_MATCHED_CITY_COLUMNS"],
            bludot_field_mapping=schema["BUSINESS_MATCHED_BLUDOT_COLUMNS"], original_record_list=schema["COLUMNS_LIST"],
            updated_record_list=schema["COLUMNS_LIST_UPDATED"], country_state_mapping=schema.get("COUNTRY_STATE_MAPPING", {}),
            method=True, city_records=city_records, uuid_for_additional=uuid_for_additional, source_type="user")

        business_matched.to_excel(str(final_excel / "Business_Matched_Records.xlsx"), index=False)
        add_business_matched.to_excel(str(final_excel / "Additional_Business_Matched_Records.xlsx"), index=False)
        stats["business_matched"], stats["business_additional"] = len(business_matched), len(add_business_matched)

    # CUSTOM RECORDS
    custom_default_value_indexes, add_custom_default_value_indexes = [], []
    if schema.get("BUSINESS_CUSTOM_MATCHED_CITY_COLUMNS") or schema.get("NEW_FIELDS"):
        custom_matched_records, custom_default_value_indexes = get_custom_matched_records(
            dataset=total_match_records, city_field_mapping=schema.get("BUSINESS_CUSTOM_MATCHED_CITY_COLUMNS", []),
            bludot_field_mapping=schema.get("BUSINESS_CUSTOM_MATCHED_BLUDOT_COLUMNS", []), method=False, uuid_for_additional='',
            new_fields=schema.get("NEW_FIELDS", []), raw_sheet=raw_sheet, city_records=city_records, filename_output=str(final_excel))
            
        add_custom_matched_records, add_custom_default_value_indexes = get_custom_matched_records(
            dataset=additional_city_match_records, city_field_mapping=schema.get("BUSINESS_CUSTOM_MATCHED_CITY_COLUMNS", []),
            bludot_field_mapping=schema.get("BUSINESS_CUSTOM_MATCHED_BLUDOT_COLUMNS", []), method=True,
            uuid_for_additional=uuid_for_additional, raw_sheet=raw_sheet, new_fields=schema.get("NEW_FIELDS", []), city_records=city_records, filename_output=str(final_excel))
            
        custom_matched_records.to_excel(str(final_excel / "Custom_Matched_Records.xlsx"), index=False)
        add_custom_matched_records.to_excel(str(final_excel / "Additional_Custom_Matched_Records.xlsx"), index=False)
        stats["custom_matched"] = len(custom_matched_records)

    # CONTACT RECORDS
    if schema.get("CONTACT_MATCHED_CITY_COLUMNS"):
        contact_matched = get_contact_matched_records(
            dataset=total_match_records, city_matched_records=schema.get("CONTACT_MATCHED_CITY_COLUMNS", []),
            bludot_matched_records=schema.get("CONTACT_MATCHED_BLUDOT_COLUMNS", []), uuid_for_additional='', method=False)
            
        add_contact_matched = get_contact_matched_records(
            dataset=additional_city_match_records, city_matched_records=schema.get("CONTACT_MATCHED_CITY_COLUMNS", []),
            bludot_matched_records=schema.get("CONTACT_MATCHED_BLUDOT_COLUMNS", []), uuid_for_additional=uuid_for_additional, method=True)
            
        contact_matched.to_excel(str(final_excel / "Contact_Matched_Records.xlsx"), index=False)
        add_contact_matched.to_excel(str(final_excel / "Additional_Contact_Matched_Records.xlsx"), index=False)
        stats["contact_matched"] = len(contact_matched)

    # WRITE EXCEL
    with pd.ExcelWriter(str(final_output / f"{city_name}_Business_Matched_Records.xlsx")) as writer:
        if schema.get("BUSINESS_MATCHED_CITY_COLUMNS"): business_matched.to_excel(writer, sheet_name="Business_Matched_Records", index=False)
        if custom_matched_records is not None: custom_matched_records.to_excel(writer, sheet_name="Custom_Matched_Records", index=False)
        if schema.get("CONTACT_MATCHED_CITY_COLUMNS"): contact_matched.to_excel(writer, sheet_name="Contact_Matched_Records", index=False)

    with pd.ExcelWriter(str(final_output / f"Additional_Matched_Records_Of_{city_name}.xlsx")) as writer:
        if schema.get("BUSINESS_MATCHED_CITY_COLUMNS"): add_business_matched.to_excel(writer, sheet_name="Additional_Business_Matched_Rec", index=False)
        if custom_matched_records is not None: add_custom_matched_records.to_excel(writer, sheet_name="Additional_Custom_Matched_Rec", index=False)
        if schema.get("CONTACT_MATCHED_CITY_COLUMNS"): add_contact_matched.to_excel(writer, sheet_name="Additional_Contact_Matched_Rec", index=False)

    if custom_matched_records is not None:
        wb1 = load_workbook(str(final_output / f"{city_name}_Business_Matched_Records.xlsx"))
        format_custom_subsheet(wb1, "Custom_Matched_Records", custom_default_value_indexes)
        wb1["Custom_Matched_Records"].protection.sheet = False
        wb1.save(str(final_output / f"{city_name}_Business_Matched_Records.xlsx"))
        
        wb2 = load_workbook(str(final_output / f"Additional_Matched_Records_Of_{city_name}.xlsx"))
        format_custom_subsheet(wb2, "Additional_Custom_Matched_Rec", add_custom_default_value_indexes)
        wb2["Additional_Custom_Matched_Rec"].protection.sheet = False
        wb2.save(str(final_output / f"Additional_Matched_Records_Of_{city_name}.xlsx"))

    city.output_file_path = str(final_output / f"{city_name}_Business_Matched_Records.xlsx")
    db.commit()

    logger.info(f"Step 5 complete: {stats}")
    return stats