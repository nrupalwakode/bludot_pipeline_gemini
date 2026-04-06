import pandas as pd
import re
import os
import glob
import datetime
import json
import logging
from pathlib import Path
from openpyxl import load_workbook
from typing import Dict, List, Optional, Tuple, Any

# Ensure the backend directory is in the path for imports
import sys
backend_dir = str(Path(__file__).resolve().parent.parent)
if backend_dir not in sys.path:
    sys.path.append(backend_dir)

from src.contact_formatting import format_contact_data, clean_column_names
from sqlalchemy.orm import Session
from ..db.models import City

logger = logging.getLogger(__name__)

# ================ HELPER FUNCTIONS ================

def is_hardcoded_value(value: Any) -> bool:
    return isinstance(value, str) and value.startswith('[') and value.endswith(']')

def extract_hardcoded_value(value: Any) -> str:
    if is_hardcoded_value(value):
        return value[1:-1]
    return value

def _get_actual_col(columns, expected_name):
    """Aggressive Fuzzy Matcher: ignores case, spaces, and handles fallbacks."""
    if not expected_name: return None
    if expected_name in columns: return expected_name
    if f"{expected_name}_1" in columns: return f"{expected_name}_1"
    
    clean_expected = str(expected_name).strip().lower()
    for col in columns:
        clean_col = str(col).strip().lower()
        if clean_col == clean_expected or clean_col == f"{clean_expected}_1": return col
        
    clean_expected_no_space = clean_expected.replace(' ', '').replace('_', '')
    for col in columns:
        clean_col_no_space = str(col).strip().lower().replace(' ', '').replace('_', '')
        if clean_col_no_space == clean_expected_no_space or clean_col_no_space == f"{clean_expected_no_space}1": return col
        
    fallbacks = []
    if "email" in clean_expected: fallbacks = ["email", "mainemail", "accountemail", "contactemail"]
    elif "phone" in clean_expected: fallbacks = ["phone", "phonenumber", "mainaccountphone", "accountphone", "contactphone"]
    elif "name" in clean_expected and "business" not in clean_expected and "account" not in clean_expected:
        fallbacks = ["name", "contactname", "personname"]
        
    for col in columns:
        clean_col_no_space = str(col).strip().lower().replace(' ', '').replace('_', '')
        if clean_col_no_space in fallbacks or f"{clean_col_no_space}1" in fallbacks:
            return col
    return None

def _get_val(row: pd.Series, expected_col: str) -> str:
    actual_col = _get_actual_col(row.index, expected_col)
    if actual_col and pd.notna(row[actual_col]):
        val = str(row[actual_col]).strip()
        if val.lower() not in ('', 'nan', 'nat', 'none', '-'): return val
    return ""

def _get_full_name(row: pd.Series, mapped_col: Any) -> str:
    """Smart Concatenator: Handles both auto-merges and explicitly mapped arrays"""
    if not mapped_col: return ""
    
    # Handle the array coming from the UI Checkboxes!
    if isinstance(mapped_col, list):
        parts = []
        for col in mapped_col:
            val = _get_val(row, col)
            if val and val != "-": parts.append(val)
        return " ".join(parts).strip()
    
    val1 = _get_val(row, mapped_col)
    clean_mapped = str(mapped_col).strip().lower()
    
    if "first" in clean_mapped:
        mid_col = str(mapped_col).replace("First", "Middle").replace("first", "middle")
        last_col = str(mapped_col).replace("First", "Last").replace("first", "last")
        
        if "cntct" in clean_mapped and "first" in clean_mapped:
            prefix = str(mapped_col).split("-")[0]
            mid_col = f"{prefix}-Mid-Other-Name"
            last_col = f"{prefix}-Last-Name"
            
        val2 = _get_val(row, mid_col)
        val3 = _get_val(row, last_col)
        
        parts = [p for p in [val1, val2, val3] if p and p != "-"]
        if parts:
            return " ".join(parts).strip()
            
    return val1 if val1 else ""

def find_dynamic_columns(columns: List[str]) -> Dict[str, Dict[int, str]]:
    dynamic_columns = {}
    pattern = r'^(.+)_(\d+)$'
    for col in columns:
        match = re.match(pattern, col)
        if match:
            base_name = match.group(1)
            suffix = int(match.group(2))
            if base_name not in dynamic_columns:
                dynamic_columns[base_name] = {}
            dynamic_columns[base_name][suffix] = col
    return dynamic_columns

def generate_id_column(df: pd.DataFrame, city_name: str, reference_date: str = None) -> pd.Series:
    city_prefix = city_name.split('_')[0][:3].upper()
    if reference_date:
        date_suffix = reference_date
    else:
        date_suffix = datetime.now().strftime("%d%m%Y")
    
    total_rows = len(df)
    padding = len(str(total_rows))
    ids = []
    for i in range(1, total_rows + 1):
        sequential_num = str(i).zfill(padding)
        id_value = f"{city_prefix}{date_suffix}{sequential_num}"
        ids.append(id_value)
    return pd.Series(ids)

def extract_date_from_city_name(city_name: str) -> str:
    try:
        parts = city_name.split('_')
        if len(parts) >= 3 and not parts[-4].isdigit():
            day = parts[-3].zfill(2)
            month = parts[-2].zfill(2)
            year = parts[-1]
            if (day.isdigit() and month.isdigit() and year.isdigit() and 
                len(day) == 2 and len(month) == 2 and len(year) == 4):
                return f"{year}{month}{day}"
        elif len(parts) >= 3:
            day = parts[-4].zfill(2)
            month = parts[-3].zfill(2)
            year = parts[-2]
            if (day.isdigit() and month.isdigit() and year.isdigit() and 
                len(day) == 2 and len(month) == 2 and len(year) == 4):
                return f"{year}{month}{day}"
    except:
        pass
    return None

def find_business_matched_columns(columns: List[str]) -> Dict[str, Any]:
    name_columns = []
    title_columns = []
    roles_columns = []
    
    name_pattern = r'^Name_y(?:\.(\d+))?$'
    title_pattern = r'^Title(?:\.(\d+))?$'
    roles_pattern = r'^Roles(?:\.(\d+))?$'
    
    name_positions = []
    for i, col in enumerate(columns):
        match = re.match(name_pattern, col)
        if match:
            suffix = int(match.group(1)) if match.group(1) else 0
            name_positions.append((suffix, i, col))
    
    name_positions.sort(key=lambda x: x[0])
    name_columns = [col for _, _, col in name_positions]
    
    title_positions = []
    for i, col in enumerate(columns):
        match = re.match(title_pattern, col)
        if match:
            suffix = int(match.group(1)) if match.group(1) else 0
            title_positions.append((suffix, i, col))
    
    title_positions.sort(key=lambda x: x[0])
    title_columns = [col for _, _, col in title_positions]
    
    roles_positions = []
    for i, col in enumerate(columns):
        match = re.match(roles_pattern, col)
        if match:
            suffix = int(match.group(1)) if match.group(1) else 0
            roles_positions.append((suffix, i, col))
    
    roles_positions.sort(key=lambda x: x[0])
    roles_columns = [col for _, _, col in roles_positions]
    
    contact_pattern = r'^(Contact|Contact_type|Type)(?:\.(\d+))?$'
    contact_positions = []
    for i, col in enumerate(columns):
        match = re.match(contact_pattern, col)
        if match:
            base_type = match.group(1)
            suffix = int(match.group(2)) if match.group(2) else 0
            contact_positions.append((i, suffix, col, base_type))
    
    contact_positions.sort(key=lambda x: x[0])
    contact_sets = {}
    
    for pos, suffix, col, base_type in contact_positions:
        current_name_idx = 0
        for i, (_, name_pos, _) in enumerate(name_positions):
            if name_pos < pos:
                current_name_idx = i
            else:
                break
        
        name_key = str(current_name_idx)
        if name_key not in contact_sets:
            contact_sets[name_key] = []
        contact_sets[name_key].append(col)
    
    name_to_contacts = {}
    for i, (_, _, name_col) in enumerate(name_positions):
        name_key = str(i)
        name_to_contacts[i] = {
            'name_col': name_col,
            'title_col': title_columns[i] if i < len(title_columns) else '',
            'roles_col': roles_columns[i] if i < len(roles_columns) else '',
            'contact_cols': contact_sets.get(name_key, [])
        }
    
    return {
        'name_columns': name_columns,
        'title_columns': title_columns,
        'roles_columns': roles_columns,
        'contact_sets': contact_sets,
        'name_to_contacts': name_to_contacts
    }

def process_contacts(df: pd.DataFrame, id_df_path: str, sheet_type: str, contact_config: Dict, city_name: str) -> pd.DataFrame:
    index_col = df.columns[0] if 'index' in df.columns[0].lower() else None
    
    dynamic_columns = find_dynamic_columns(df.columns)
    max_suffix = 0
    for base_dict in dynamic_columns.values():
        if base_dict and max(base_dict.keys()) > max_suffix:
            max_suffix = max(base_dict.keys())
            
    if max_suffix == 0:
        max_suffix = 1
    
    output_columns = ['ID']
    for i in range(1, len(contact_config) + 1):
        prefix = "" if i == 1 else f"{i}_"
        output_columns.extend([
            f"{prefix}Name", f"{prefix}Title", f"{prefix}Roles", 
            f"{prefix}Contact", f"{prefix}Contact_type", f"{prefix}Type"
        ])
    
    if index_col:
        output_columns = [index_col] + output_columns    

    output_df = pd.DataFrame(columns=output_columns)
    
    if os.path.exists(id_df_path):
        id_df = pd.read_excel(id_df_path)
        uuid_col = None
        for col in id_df.columns:
            if col.lower() == 'id':
                uuid_col = col
                break
        has_uuid = uuid_col is not None
    else:
        has_uuid = False

    if has_uuid:
        if 'ID' not in output_columns:
            output_columns.insert(0, 'ID')
        id_values = id_df[uuid_col].tolist()
    else:
        city_date = extract_date_from_city_name(city_name)
        if city_date:
            id_values = generate_id_column(df, city_name, city_date).tolist()
        else:
            id_values = generate_id_column(df, city_name).tolist()
            
    blank_placeholders = {}
    blank_counter = 1

    for idx, row in df.iterrows():
        output_row = {}
        if index_col:
            output_row[index_col] = row[index_col]
            
        output_row['ID'] = id_values[idx] if idx < len(id_values) else ""
        
        for col in output_columns:
            if col != index_col and col != 'ID':
                output_row[col] = ""
        
        # --- ROBUST FUZZY MATCHER EXTRACTION ---
        all_persons = {}
        for suffix in range(1, max_suffix + 1):
            for config_val in contact_config.values():
                person_base = config_val.get("person_col", "")
                if not person_base or is_hardcoded_value(person_base): continue
                
                if isinstance(person_base, list):
                    p_vals = []
                    for pb in person_base:
                        suff_col = _get_actual_col(df.columns, f"{pb}_{suffix}")
                        if suff_col:
                            val = _get_val(row, suff_col)
                            if val: p_vals.append(val)
                        elif suffix == 1:
                            val = _get_val(row, pb)
                            if val: p_vals.append(val)
                    p_val = " ".join(p_vals).strip()
                    if p_val:
                        all_persons[(tuple(person_base), suffix)] = p_val
                else:
                    suff_col = _get_actual_col(df.columns, f"{person_base}_{suffix}")
                    if suff_col:
                        p_val = _get_val(row, suff_col)
                        if p_val: all_persons[(person_base, suffix)] = p_val
                    elif suffix == 1:
                        p_val = _get_full_name(row, person_base)
                        if p_val: all_persons[(person_base, suffix)] = p_val
        
        all_contacts = []
        processed_persons = set()
        
        for suffix in range(1, max_suffix + 1):
            for config_idx, (contact_group, config) in enumerate(contact_config.items(), 1):
                person_col_base = config.get("person_col", "")
                title_col_base = config.get("title_col", "")
                roles_col_base = config.get("roles_col", "")
                contact_col_base = config.get("contact_col", "")
                contact_type_base = config.get("contact_type", "")
                type_base = config.get("type", "")
                
                # Person
                person_val = ""
                if is_hardcoded_value(person_col_base):
                    person_val = extract_hardcoded_value(person_col_base)
                elif isinstance(person_col_base, list):
                    p_vals = []
                    for pb in person_col_base:
                        suff_col = _get_actual_col(df.columns, f"{pb}_{suffix}")
                        if suff_col:
                            val = _get_val(row, suff_col)
                            if val: p_vals.append(val)
                        elif suffix == 1:
                            val = _get_val(row, pb)
                            if val: p_vals.append(val)
                    person_val = " ".join(p_vals).strip()
                else:
                    suff_col = _get_actual_col(df.columns, f"{person_col_base}_{suffix}")
                    if suff_col: person_val = _get_val(row, suff_col)
                    elif suffix == 1: person_val = _get_full_name(row, person_col_base)
                        
                # Contact
                contact_val = ""
                if is_hardcoded_value(contact_col_base):
                    contact_val = extract_hardcoded_value(contact_col_base)
                else:
                    suff_col = _get_actual_col(df.columns, f"{contact_col_base}_{suffix}")
                    if suff_col: contact_val = _get_val(row, suff_col)
                    elif suffix == 1: contact_val = _get_val(row, contact_col_base)
                        
                # Title
                title_val = ""
                if is_hardcoded_value(title_col_base):
                    title_val = extract_hardcoded_value(title_col_base)
                else:
                    suff_col = _get_actual_col(df.columns, f"{title_col_base}_{suffix}")
                    if suff_col: title_val = _get_val(row, suff_col)
                    elif suffix == 1: title_val = _get_val(row, title_col_base)
                        
                # Roles
                role_val = ""
                if is_hardcoded_value(roles_col_base):
                    role_val = extract_hardcoded_value(roles_col_base)
                else:
                    suff_col = _get_actual_col(df.columns, f"{roles_col_base}_{suffix}")
                    if suff_col: role_val = _get_val(row, suff_col)
                    elif suffix == 1: role_val = _get_val(row, roles_col_base)

                # Contact Type
                contact_type_val = ""
                if is_hardcoded_value(contact_type_base):
                    contact_type_val = extract_hardcoded_value(contact_type_base)
                else:
                    suff_col = _get_actual_col(df.columns, f"{contact_type_base}_{suffix}")
                    if suff_col: contact_type_val = _get_val(row, suff_col)
                    elif suffix == 1: contact_type_val = _get_val(row, contact_type_base)

                # Type
                type_val = ""
                if is_hardcoded_value(type_base):
                    type_val = extract_hardcoded_value(type_base)
                else:
                    suff_col = _get_actual_col(df.columns, f"{type_base}_{suffix}")
                    if suff_col: type_val = _get_val(row, suff_col)
                    elif suffix == 1: type_val = _get_val(row, type_base)

                if contact_val and not person_val:
                    pb_key = tuple(person_col_base) if isinstance(person_col_base, list) else person_col_base
                    placeholder_key = f"{pb_key}_{suffix}"
                    if placeholder_key not in blank_placeholders:
                        blank_placeholders[placeholder_key] = f"BLK_{blank_counter}"
                        blank_counter += 1
                    person_val = blank_placeholders[placeholder_key]
                
                if contact_val or person_val:
                    final_contact_type = contact_type_val if contact_val else ""
                    final_type_val = type_val if contact_val else ""
                    all_contacts.append({
                        "name": person_val,
                        "title": title_val,
                        "roles": role_val,
                        "contact": contact_val,
                        "contact_type": final_contact_type,
                        "type": final_type_val
                    })
                    
                    if person_val and person_val != "-":
                        processed_persons.add(person_val)
        
        people_without_contacts = set()
        for (person_base, suffix), person_name in all_persons.items():
            if person_name and person_name not in processed_persons and person_name not in people_without_contacts:
                all_contacts.append({
                    "name": person_name,
                    "title": "",
                    "roles": "",
                    "contact": "",
                    "contact_type": "",
                    "type": ""
                })
                people_without_contacts.add(person_name)
        
        next_contact_index = 1
        
        for i, contact in enumerate(all_contacts):
            slot_prefix = "" if next_contact_index == 1 else f"{next_contact_index}_"
            
            required_cols = [
                f"{slot_prefix}Name", f"{slot_prefix}Title", f"{slot_prefix}Roles",
                f"{slot_prefix}Contact", f"{slot_prefix}Contact_type", f"{slot_prefix}Type"
            ]
            
            for col in required_cols:
                if col not in output_row:
                    output_row[col] = ""
                if col not in output_columns:
                    output_columns.append(col)
            
            output_row[f"{slot_prefix}Name"] = contact["name"]
            output_row[f"{slot_prefix}Title"] = contact["title"]
            output_row[f"{slot_prefix}Roles"] = contact["roles"]
            output_row[f"{slot_prefix}Contact"] = contact["contact"]
            output_row[f"{slot_prefix}Contact_type"] = contact["contact_type"]
            output_row[f"{slot_prefix}Type"] = contact["type"]
            
            next_contact_index += 1
        
        if sheet_type.lower() == "business_matched":
            business_matched_cols = find_business_matched_columns(df.columns)
            business_contacts = []
            name_to_contacts = business_matched_cols.get('name_to_contacts', {})
            
            for name_idx in sorted(name_to_contacts.keys()):
                name_info = name_to_contacts[name_idx]
                name_col = name_info['name_col']
                title_col = name_info['title_col']
                roles_col = name_info['roles_col']
                contact_cols = name_info['contact_cols']
                
                name_val = row[name_col] if name_col in df.columns and pd.notna(row[name_col]) else ""
                title_val = row[title_col] if title_col and title_col in df.columns and pd.notna(row[title_col]) else ""
                roles_val = row[roles_col] if roles_col and roles_col in df.columns and pd.notna(row[roles_col]) else ""
                
                contact_dict = {}
                
                for col in contact_cols:
                    if col.startswith('Contact') and not col.startswith('Contact_type'):
                        if col == 'Contact':
                            suffix = 0
                        else:
                            suffix_match = re.search(r'\.(\d+)$', col)
                            suffix = int(suffix_match.group(1)) if suffix_match else 0
                        
                        if suffix not in contact_dict:
                            contact_dict[suffix] = {'contact': '', 'contact_type': '', 'type': ''}
                        contact_dict[suffix]['contact'] = col
                        
                    elif col.startswith('Contact_type'):
                        if col == 'Contact_type':
                            suffix = 0
                        else:
                            suffix_match = re.search(r'\.(\d+)$', col)
                            suffix = int(suffix_match.group(1)) if suffix_match else 0
                        
                        if suffix not in contact_dict:
                            contact_dict[suffix] = {'contact': '', 'contact_type': '', 'type': ''}
                        contact_dict[suffix]['contact_type'] = col
                        
                    elif col.startswith('Type'):
                        if col == 'Type':
                            suffix = 0
                        else:
                            suffix_match = re.search(r'\.(\d+)$', col)
                            suffix = int(suffix_match.group(1)) if suffix_match else 0
                        
                        if suffix not in contact_dict:
                            contact_dict[suffix] = {'contact': '', 'contact_type': '', 'type': ''}
                        contact_dict[suffix]['type'] = col
                
                has_any_contact = False
                for suffix in sorted(contact_dict.keys()):
                    triplet = contact_dict[suffix]
                    
                    contact_val = ""
                    if triplet['contact'] and triplet['contact'] in df.columns:
                        contact_val = row[triplet['contact']] if pd.notna(row[triplet['contact']]) else ""
                    
                    contact_type_val = ""
                    if triplet['contact_type'] and triplet['contact_type'] in df.columns:
                        contact_type_val = row[triplet['contact_type']] if pd.notna(row[triplet['contact_type']]) else ""
                    
                    type_val = ""
                    if triplet['type'] and triplet['type'] in df.columns:
                        type_val = row[triplet['type']] if pd.notna(row[triplet['type']]) else ""
                    
                    if contact_val:
                        has_any_contact = True
                        final_name = name_val
                        if not final_name and contact_val:
                            placeholder_key = f"business_{name_idx}_{suffix}"
                            if placeholder_key not in blank_placeholders:
                                blank_placeholders[placeholder_key] = f"BLK_{blank_counter}"
                                blank_counter += 1
                            final_name = blank_placeholders[placeholder_key]
                        
                        business_contacts.append({
                            "name": final_name,
                            "title": title_val,
                            "roles": roles_val,
                            "contact": contact_val,
                            "contact_type": contact_type_val,
                            "type": type_val
                        })
                
                if not has_any_contact and name_val:
                    business_contacts.append({
                        "name": name_val,
                        "title": title_val,
                        "roles": roles_val,
                        "contact": "",
                        "contact_type": "",
                        "type": ""
                    })
            
            for contact in business_contacts:
                slot_prefix = "" if next_contact_index == 1 else f"{next_contact_index}_"
                
                required_cols = [
                    f"{slot_prefix}Name", f"{slot_prefix}Title", f"{slot_prefix}Roles",
                    f"{slot_prefix}Contact", f"{slot_prefix}Contact_type", f"{slot_prefix}Type"
                ]
                
                for col in required_cols:
                    if col not in output_row:
                        output_row[col] = ""
                    if col not in output_columns:
                        output_columns.append(col)
                
                output_row[f"{slot_prefix}Name"] = contact["name"]
                output_row[f"{slot_prefix}Title"] = contact["title"]
                output_row[f"{slot_prefix}Roles"] = contact["roles"]
                output_row[f"{slot_prefix}Contact"] = contact["contact"]
                output_row[f"{slot_prefix}Contact_type"] = contact["contact_type"]
                output_row[f"{slot_prefix}Type"] = contact["type"]
                
                next_contact_index += 1

        output_df = pd.concat([output_df, pd.DataFrame([output_row])], ignore_index=True)
    
    return output_df[output_columns]

def process_file(input_file_path: str, id_df_path: str, output_file_path: str, sheet_type: str, contact_config: Dict, city_name: str) -> int:
    logger.info(f"Processing file: {input_file_path} as {sheet_type}")
    
    if not os.path.exists(input_file_path):
        logger.error(f"Error: Input file '{input_file_path}' not found.")
        return 0
    
    try:
        if input_file_path.endswith('.csv'):
            df = pd.read_csv(input_file_path)
        else:
            df = pd.read_excel(input_file_path)
        
        output_df = process_contacts(df, id_df_path, sheet_type, contact_config, city_name)
        
        os.makedirs(os.path.dirname(output_file_path), exist_ok=True)
        
        processed_df = format_contact_data(output_df, max_workers=None)
        processed_df = clean_column_names(processed_df)
        processed_df = processed_df.replace(to_replace=r'^BLK_\d+$', value='-', regex=True)

        expected_cols = ["ID", "Name", "Title", "Roles", "Contact", "Contact_type", "Type"]
        for req_col in expected_cols:
            if req_col not in processed_df.columns:
                processed_df[req_col] = ""

        # CRITICAL FIX: Make columns temporarily unique to prevent 'DataFrame has no attribute str' crashes
        original_cols = processed_df.columns.tolist()
        processed_df.columns = [f"{c}_{i}" for i, c in enumerate(original_cols)]

        # STRICT FORMATTING CLEANUP (Blank vs "-" Rules)
        for col, orig_col in zip(processed_df.columns, original_cols):
            if orig_col == 'ID': continue
            
            processed_df[col] = processed_df[col].fillna("").astype(str).str.strip()
            processed_df[col] = processed_df[col].replace(["nan", "NaN", "None", "<NA>", "nan.0"], "")
            
            if 'Name' in orig_col:
                processed_df[col] = processed_df[col].replace("", "-")
            else:
                processed_df[col] = processed_df[col].replace("-", "")
                
        # Restore original duplicate columns
        processed_df.columns = original_cols
        
        processed_df.to_csv(output_file_path, index=False)
        return len(processed_df)
        
    except Exception as e:
        logger.error(f"Error processing file: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return 0

def copy_csv_to_excel(csv_path, excel_path, sheet_name):
    if not os.path.exists(csv_path) or not os.path.exists(excel_path):
        return

    with open(csv_path, "r", encoding="utf-8") as f:
        header_line = f.readline().strip()
    
    original_columns = header_line.split(",")

    df = pd.read_csv(csv_path, header=0)
    df.columns = original_columns

    workbook = load_workbook(excel_path)

    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    ws = workbook.create_sheet(title=sheet_name)
    ws.append(original_columns)

    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))

    workbook.save(excel_path)

# =========================================================================
# ORCHESTRATOR BRIDGE
# =========================================================================
def run_step6(city: City, db: Session, results_dir: str) -> dict:
    results_path = Path(results_dir)
    city_dir = results_path.parent
    city_name = city.name

    INPUT_FOLDER = str(results_path / "output" / "final_result")
    OUTPUT_FOLDER = str(results_path / "output" / "final_excel")
    FINAL_OUTPUT_FOLDER = str(results_path / "output" / "final_output")
    
    schema_path = city_dir / "city_schema.json"
    if not schema_path.exists():
        logger.error(f"Step 6: Missing schema at {schema_path}")
        return {"contacts_processed": 0}
        
    with open(str(schema_path)) as f:
        schema = json.load(f)
        
    contact_config = schema.get("CONTACT_CONFIG", {})
    if not contact_config:
        logger.info("Step 6: No CONTACT_CONFIG found in schema. Skipping.")
        return {"contacts_processed": 0}

    sheet_types = ["additional_records", "business_matched"]
    total_processed = 0
    
    for sheet_type in sheet_types:
        logger.info(f"Processing files for sheet type: {sheet_type}")
        
        if sheet_type == "additional_records":
            input_pattern = os.path.join(INPUT_FOLDER, f"additional_city_records_for_{city_name}.xlsx")
            id_df_pattern = os.path.join(OUTPUT_FOLDER, f"Additional_Business_Matched_Records.xlsx")
            output_path = os.path.join(OUTPUT_FOLDER, "Additional_Contact_Matched_Records.csv")
        else:
            input_pattern = os.path.join(INPUT_FOLDER, f"final_matched_records_for_{city_name}.xlsx")
            id_df_pattern = os.path.join(OUTPUT_FOLDER, "Business_Matched_Records.xlsx")
            output_path = os.path.join(OUTPUT_FOLDER, "Contact_Matched_Records.csv")
        
        input_files = glob.glob(input_pattern)
        if not input_files:
            continue
            
        for input_file in input_files:
            rows_processed = process_file(input_file, id_df_pattern, output_path, sheet_type, contact_config, city_name)
            total_processed += rows_processed
    
    csv_additional_contact = os.path.join(OUTPUT_FOLDER, "Additional_Contact_Matched_Records.csv")
    csv_contact = os.path.join(OUTPUT_FOLDER, "Contact_Matched_Records.csv")
    
    excel_additional = os.path.join(FINAL_OUTPUT_FOLDER, f"Additional_Matched_Records_Of_{city_name}.xlsx")
    excel_business = os.path.join(FINAL_OUTPUT_FOLDER, f"{city_name}_Business_Matched_Records.xlsx")

    copy_csv_to_excel(csv_additional_contact, excel_additional, "Additional_Contact_Matched_Rec")
    copy_csv_to_excel(csv_contact, excel_business, "Contact_Matched_Records")

    logger.info("Step 6 complete. All sheets created successfully.")
    return {"contacts_processed": total_processed}