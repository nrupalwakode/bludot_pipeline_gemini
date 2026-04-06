import os
import json
import math
import numpy as np
import pandas as pd
import string
import warnings
from datetime import datetime
from openpyxl.styles import Alignment,Border, Side, Font

warnings.filterwarnings("ignore")

def uuid_sequence_for_additional(business_matched_records,uuid_for_additional):
    additional_uuid=[]
    max_row_count_len=len(str(business_matched_records.shape[0]))
    for i in range(1,business_matched_records.shape[0]+1):
            x=uuid_for_additional+'0'*(max_row_count_len-len(str(i)))+str(i)
            additional_uuid.append(x)
    return additional_uuid

def single_column_conversion(dataset,columns_name):
    output_string =[]
    dataset.fillna('',inplace=True)
       
    for mylist in dataset[columns_name].values:
        # print(dataset[columns_name])
        string_output = ''
        len_mylist = len(mylist)
        # if (len(mylist)!=0): # updated line
        if (mylist[0]!='' and mylist[0]!='-'):
                #print(mylist[0],0)
                    string_output = mylist[0]
        else:
            for numbers in range(1,len_mylist):
                if mylist[numbers]!='':
                    #print(mylist[numbers],numbers)
                    string_output = mylist[numbers]
                    break  
                else:
                    string_output = ''
        output_string.append(string_output)
    # print(output_string)
    return output_string

def data_reconstruction(dataset,columns_name,name):
    main_contact,index_values =[],[]
    # print(columns_name)
    combine_details =  single_column_conversion(dataset = dataset,
                                                columns_name=columns_name)  
    
    
    for ref_value,ref_index in zip(combine_details,dataset.index):
        sample_dataset = dataset[columns_name]
        common_list_details = []
        
        for remaining_details in sample_dataset.iloc[ref_index,:].values:
            # print(k,i,"\t",type(k),type(k))

            if len(str(remaining_details))>=2:          #type casted bcs tyape error for int/float data
                if ref_value!=remaining_details:
                    common_list_details.append(remaining_details)
        main_contact.append((common_list_details,ref_index)) 
        
    values_data,index_values =zip(*main_contact)
    updated_length = []
    for max_column_count in values_data:
        if len(max_column_count) not in updated_length:
            updated_length.append(len(max_column_count))  

    #print(f"{name} columns has {max(updated_length)} sub-columns are Updated in Final Sheet")
    mydata = pd.DataFrame(values_data,columns=[f'{name}_{p}' for p in range(2,max(updated_length)+2)])

    # mydata[f'{name}_index'] = index_values
    # mydata.fillna('',inplace=True)
    if  mydata.shape[1]<=1:
        final_output = pd.DataFrame(combine_details,columns=[f'{name}'])
    else:
         final_output = pd.DataFrame(combine_details,columns=[f'{name}_{1}'])
    # print(',,,,,,,,,,',final_output)
    #print(mydata.shape,name,final_output.columns)            
    concated_output = pd.concat([final_output,mydata],axis=1)
    concated_output.fillna('',inplace=True)
    concated_output=concated_output.mask(concated_output == '')
    concated_output[name] =concated_output.apply(lambda x: ';'.join(x.dropna().astype(str).values), axis=1)
    df=pd.DataFrame(concated_output[name])
    df[f'{name}_index'] = index_values
    return  df

dropdown_datatype_columns={
    'Primary SIC Code and Sector':'Number',
    'PRIMARY SIC SECTOR':'Text(Small)',
    'Primary NAICS Code and Sector':'Number',
    'PRIMARY NAICS SECTOR':'Text(Small)'
}

def uuid_sequence_for_additional(business_matched_records,uuid_for_additional):
    additional_uuid=[]
    max_row_count_len=len(str(business_matched_records.shape[0]))
    for i in range(1,business_matched_records.shape[0]+1):
            x=uuid_for_additional+'0'*(max_row_count_len-len(str(i)))+str(i)
            additional_uuid.append(x)
    return additional_uuid

def isfloat(num):
    try:
        float(num)
        return True
    except ValueError:
        return False
    
def largest_num_list(city_df,updated_col_name,index):
    max_number_value_list=[]
    for data in city_df[updated_col_name[index]]:
        # print(data)
        data=str(data)
        data=data.replace(';',' ')
        if data!='':
            data1=data.split(" ")
            max_number_value=[]
            for i in data1:
                if i.isnumeric():
                    max_number_value.append(int(i))
                elif isfloat(i):
                    max_number_value.append(math.ceil(float(i)))
                elif not i.isalpha():
                    res = ''.join(filter(lambda x: x.isdigit(),i))
                    max_number_value.append(int(res))
            if len(max_number_value)>=1:
                max_number_value_list.append(max(max_number_value))
            else:
                max_number_value_list.append('')
        else:
            max_number_value_list.append('')
    return max_number_value_list

def earliest_year_list(city_df,updated_col_name,index):
    min_year_list=[]
    for data in city_df[updated_col_name[index]]:
        data=str(data)
        data=data.replace(',',' ').replace('-',' ').replace('/',' ').replace(':',' ').replace(';',' ')
        if data!='':
            data1=data.split(" ")
            year=[]
            for i in data1:
                if i.isnumeric():
                    if 9999>=int(i)>=1000:
                        year.append(i)
            if len(year)>=1:
                min_year_list.append(min(year))
            else:
                min_year_list.append('')
        else:
            min_year_list.append('')
    return min_year_list

def earliest_date_list(city_df,updated_col_name,index):
    earliest_date_list=[]
    for data in city_df[updated_col_name[index]]:
        data=str(data)
        data=data.replace(';',' ')
        if data!='':
            data1=data.split(" ")
            date1=[]
            for i in data1:
                date1.append(datetime.strptime(i,'%m/%d/%Y'))
            if len(date1)==1:
                earliest_date_list.append(date1[0].strftime("%m/%d/%Y"))
            else:
                x=date1[0]
                for j in date1[1:]:
                    if x >j:
                        x=j
                earliest_date_list.append(x.strftime("%m/%d/%Y"))
        else:
            earliest_date_list.append('') 
    return earliest_date_list
    

def datatype_by_column_data(dataset,new_field_col):
    column_data = dataset[f'{new_field_col}_1']
    # Check if the column contains numerical values
    if pd.to_numeric(column_data, errors='coerce').notnull().all():
        col_datatype='Number'

    # Check if the column contains year values
    elif column_data.astype(str).str.match(r'^\d{4}$').all():
        col_datatype='Year'

    # Check if the column contains date values
    elif pd.to_datetime(column_data, errors='coerce').notnull().all():
        col_datatype='Date'

    # If none of the above, assume it's a text column
    else:
        # Calculate the length of each value in the column
        value_lengths = column_data.astype(str).str.len()

        # Find the maximum length
        max_length = value_lengths.max()
        if max_length<256:
            col_datatype='Text(Small)'
        elif max_length<4096:
            col_datatype='Text(Medium)'
        elif max_length<65536:
            col_datatype='Text(Large)'
            
    return col_datatype

def get_custom_matched_records(dataset,city_field_mapping,bludot_field_mapping,method,uuid_for_additional,new_fields,raw_sheet,city_records,filename_output):
    if method==True:
        dataset['ID']=uuid_sequence_for_additional(dataset,uuid_for_additional)
        if len(bludot_field_mapping)!=0:
            for column in bludot_field_mapping:
                dataset[column] = ''
                
        ouput_sheet_name=os.path.join(filename_output,'Additional_Custom_Matched_Records.xlsx')
    else:
        ouput_sheet_name=os.path.join(filename_output,'Custom_Matched_Records.xlsx')
    datatype=[]
#     column_format=[]
    new_field_datatype=[]
    updated_col_name=[]
    custom_column_name=[]
    default_or_user_value=[]
    city_columns=[]
    city_df=pd.DataFrame()
    custom_output_sheet=pd.DataFrame()
    for col in (city_field_mapping):
        col_duplicates=[]
        for final_sheet_columns in dataset.columns:
            if final_sheet_columns.startswith(f'{col}_'):
                col_duplicates.append(final_sheet_columns)
        df=dataset.loc[:,col_duplicates]
        df.fillna('',inplace=True)
        df=df.mask(df == '')
        city_df[col] =df.apply(lambda x: ';'.join(set(x.dropna().astype(str).values)), axis=1)

    if len(bludot_field_mapping)!=0:
        raw_custom_sheet=pd.read_excel(raw_sheet,sheet_name='Custom Data')
        bludot_custom_columns=list(raw_custom_sheet.columns)
        for list_index,custom_col_name in enumerate(bludot_field_mapping):
            col_index=bludot_custom_columns.index(custom_col_name)
            col_datatype=raw_custom_sheet.iloc[0,col_index]
            if col_datatype=='Dropdown':
                if custom_col_name in dropdown_datatype_columns:
                    col_datatype=dropdown_datatype_columns[custom_col_name]
                else:
                    col_datatype=datatype_by_column_data(dataset,city_field_mapping[list_index])
                    
            if raw_custom_sheet.iloc[1,col_index]=='Default Value':
#                 column_format.append('Two Column Format')
                if city_field_mapping.count(city_field_mapping[list_index])==1:
                    updated_col_name.extend((custom_col_name,city_field_mapping[list_index]))
                    custom_column_name.extend((custom_col_name,city_field_mapping[list_index]))
                else:
                    dataset[f'{city_field_mapping[list_index]}_city']=dataset[city_field_mapping[list_index]]
                    city_df[f'{city_field_mapping[list_index]}_city']=city_df[city_field_mapping[list_index]]
                    updated_col_name.extend((custom_col_name,f'{city_field_mapping[list_index]}_city'))
                    custom_column_name.extend((custom_col_name,f'{city_field_mapping[list_index]}_city'))
                datatype.append(col_datatype)
                datatype.append(col_datatype)
                default_or_user_value.extend(('Default','User Defined'))
            else:
#                 column_format.append('One Column Format')
                updated_col_name.append(city_field_mapping[list_index])
                custom_column_name.append(custom_col_name)
                datatype.append(col_datatype)
                default_or_user_value.append('User Defined')
    
    if len(new_fields):
        for new_field_col in new_fields:
            col_datatype= datatype_by_column_data(dataset,new_field_col) 
            new_field_datatype.append(col_datatype)
            
    for index,col in enumerate(default_or_user_value):
        if col=='Default':
            custom_output_sheet[updated_col_name[index]]=dataset[updated_col_name[index]]
        else:
            if datatype[index]=='Number':
                custom_output_sheet[updated_col_name[index]]=largest_num_list(city_df,updated_col_name,index)
                
            elif datatype[index]=='Year':
                custom_output_sheet[updated_col_name[index]]=earliest_year_list(city_df,updated_col_name,index)
                
            elif datatype[index]=='Date':
                custom_output_sheet[updated_col_name[index]]=earliest_date_list(city_df,updated_col_name,index)
                
            else:
                custom_output_sheet[updated_col_name[index]]=city_df[updated_col_name[index]]
                
            if updated_col_name[index] in new_fields:
                updated_col_name[index]=updated_col_name[index]+datatype[index]
                custom_output_sheet.rename(columns={custom_output_sheet.columns[index]: updated_col_name[index]+datatype[index]}, inplace=True)

    new_fields_df=pd.DataFrame()
    for col in (new_fields):
        col_duplicates=[]
        for final_sheet_columns in dataset.columns:
            if final_sheet_columns.startswith(f'{col}_'):
                col_duplicates.append(final_sheet_columns)

        new_fields_col=dataset.loc[:,col_duplicates]
        new_fields_col.fillna('',inplace=True)
        new_fields_col=new_fields_col.mask(new_fields_col == '')
        new_fields_df[col] =new_fields_col.apply(lambda x: ';'.join(set(x.dropna().astype(str).values)), axis=1)
        
    for index,col_datatype in enumerate(new_field_datatype):
        if col_datatype=='Number':
            custom_output_sheet[new_fields[index]]=largest_num_list(new_fields_df,new_fields,index)
            
        elif col_datatype=='Year':
            custom_output_sheet[new_fields[index]]=earliest_year_list(new_fields_df,new_fields,index)

        elif col_datatype=='Date':
            custom_output_sheet[new_fields[index]]=earliest_date_list(new_fields_df,new_fields,index)

        else:
            custom_output_sheet[new_fields[index]]=new_fields_df[new_fields[index]]

        updated_col_name.append(new_fields[index])
        custom_column_name.append(new_fields[index])
        datatype.append(col_datatype)
        default_or_user_value.append('User Defined')
    custom_output_sheet.columns=custom_column_name
    custom_output_sheet.loc[-1] = default_or_user_value
    custom_output_sheet = custom_output_sheet.sort_index().reset_index(drop=True)
    
    custom_output_sheet.loc[-1] = datatype
    custom_output_sheet = custom_output_sheet.sort_index().reset_index(drop=True)
    
    custom_output_sheet.insert(0,'Custom Data Name',['Custom Data Type','ID']+dataset['ID'].tolist())
    default_values=[index for index,i in enumerate(default_or_user_value) if i=='Default']
    
    return (custom_output_sheet,default_values)

def format_custom_subsheet(workbook, subsheet_name, default_values):
    worksheet = workbook[subsheet_name]

    for default_value_index in default_values:
        col_start = default_value_index + 1
        col_end = default_value_index + 2
        row_start = 0
        row_end = 0

        worksheet.merge_cells(start_row=row_start + 1, start_column=col_start + 1, end_row=row_end + 1, end_column=col_end + 1)
        merged_cell = worksheet.cell(row=row_start + 1, column=col_start + 1)
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')

        row_start = 1
        row_end = 1

        worksheet.merge_cells(start_row=row_start + 1, start_column=col_start + 1, end_row=row_end + 1, end_column=col_end + 1)
        merged_cell = worksheet.cell(row=row_start + 1, column=col_start + 1)
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Define the range of rows to format (1st to 3rd row)
    start_row = 1
    end_row = 3

    # Apply formatting to each row
    for row_num in range(start_row, end_row + 1):
        row = worksheet[row_num]

        # Apply bold font to each cell in the row
        for cell in row:
            cell.font = Font(bold=True)

        # Add border to each cell in the row
        border = Border(top=Side(border_style='thin'),
                        right=Side(border_style='thin'),
                        bottom=Side(border_style='thin'),
                        left=Side(border_style='thin'))

        for cell in row:
            cell.border = border

def get_country_state(file_path, state_to_country):
    file_path=os.path.join(os.getcwd(),'src','country_state_mapping.json')
    with open(file_path, 'r') as json_file:
        country_state_mapping = json.load(json_file)
        # print(country_state_mapping)

    for country, states in country_state_mapping.items():
        for state in states.values():
            state_to_country[state] = country

    return state_to_country

def get_country(state, state_to_country):
    return state_to_country.get(state, '')

def change_sheet_format(dataset, original_record_list ,updated_record_list,country_state_mapping):
    state_to_country = get_country_state(country_state_mapping,{})

    column_mapping = {old: new for old, new in zip(original_record_list, updated_record_list)}    
    dataset.rename(columns=column_mapping, inplace=True)

    country_values = dataset['State'].apply(lambda state: get_country(state, state_to_country))
    dataset.insert(dataset.columns.get_loc('State')+1, 'Country', country_values)

    return dataset

def get_Business_Matched_Records(dataset,city_field_mapping,bludot_field_mapping,original_record_list, updated_record_list,country_state_mapping,city_records,method,uuid_for_additional,source_type='user'):
    business_matched_records = pd.DataFrame()
    if method==True: 
        update_input_details = []
        for cols_name_add in original_record_list:
            if cols_name_add not in dataset.columns:
               update_input_details.append(cols_name_add)

        for add_cols in update_input_details:
            dataset[add_cols]=''
    # fill null value with strings
    
    dataset.fillna('',inplace=True)
    
    # city columns identification and sorting based on its names
    final_mapping_report =[]
    for city_cols in city_field_mapping:
        combine_data =[]
        for original_cols in city_records.columns:
            trigger = len(original_cols)-len(city_cols)
            if (original_cols.startswith(city_cols)) & (trigger<=3):
                combine_data.append(original_cols)
        final_mapping_report.append(combine_data)
        
    mapping_field = dict()
    updated_bludot_field_mapping =[]
    
    for bludot_cols in bludot_field_mapping:
        updated_bludot_field_mapping.append(bludot_cols)
    #print(updated_bludot_field_mapping,'\n')  
    
    for key_values,key_name_bludot,key_name_original in zip(final_mapping_report,updated_bludot_field_mapping,bludot_field_mapping):
        combine_details =[]
        if len(key_name_bludot)>=3:
            #print(key_values,key_name_bludot)
            for substring in key_values:
               
                combine_details.append(substring)
                
            combine_details.append(key_name_bludot)
        else:
            #print(key_values,key_name_bludot)
            for substring in key_values:
                combine_details.append(substring)
                
        mapping_field[key_name_original] = combine_details
        
    #print mapping field    
    # print(json.dumps(mapping_field, indent=2, default=str))
    # print("Updated Field Mapping Reports- Business Records\n",json.dumps(mapping_field, indent=2, default=str))
    
    #updated DataFrame as per mapping Field.
    for updated_cols_name in original_record_list:
        if updated_cols_name in mapping_field.keys():
            updated_single_columns = single_column_conversion(dataset = dataset,
                                                              columns_name = mapping_field.get(updated_cols_name))
        else:
             updated_single_columns = dataset[updated_cols_name]

        business_matched_records[updated_cols_name] = updated_single_columns
            
    if method==True:    
    #    business_matched_records['is_business'] = 'True'
       business_matched_records['UUID'] =uuid_sequence_for_additional(business_matched_records,uuid_for_additional)


    business_matched_records = change_sheet_format(dataset= business_matched_records, original_record_list=original_record_list, updated_record_list=updated_record_list,country_state_mapping=country_state_mapping)

    business_matched_records['is_business'] = 'True'    
    business_matched_records['business_source'] = source_type

    cols = business_matched_records.columns.tolist()
    
    # 1. Pull these specific columns out of the list
    for col in ['business_source', 'DBA Name', 'Business Operational Status']:
        if col in cols:
            cols.remove(col)
            
    # 2. Put them back at the very end in your exact required order
    for col in ['business_source', 'DBA Name', 'Business Operational Status']:
        if col in business_matched_records.columns: # Only add it if it actually exists in the data
            cols.append(col)
            
    # 3. Apply the sorted column list to the dataframe
    business_matched_records = business_matched_records[cols]
    # ---------------------------------------------------------

    return business_matched_records

def isfloat(num):
    try:
        float(num)
        return True
    except ValueError:
        return False


def string_filter_for_abbreviation(strings):
    # Create a string of all punctuation characters
    punctuation_chars = string.punctuation

    # Replace each punctuation character with a space
    strings = strings.lower().translate(str.maketrans(punctuation_chars, ' ' * len(punctuation_chars)))

    # Remove any extra spaces
    strings = ' '.join(strings.split())

    return strings

def phone_number_formatting(phone_num):
    # Remove specific characters from the phone number
    phone_num = ''.join(char for char in phone_num if char.isdigit())

    return phone_num

def get_contact_matched_records(CITY_NAME,city_records,contact_city_columns,column_type,matched_records):
    city_records.fillna('', inplace=True)  # Replace any NaN values in the city_records DataFrame with empty strings
    city_records = city_records.astype(str)  # Convert all columns in the city_records DataFrame to string type
    updated_contact_city_columns = []  # Initialize an empty list to store the updated contact columns

    # Iterate over each contact column
    for contact_column in contact_city_columns:
        duplicate_column_names = []  # Initialize an empty list to store the duplicate column names for the contact column

        # Iterate over each column name in the city_records DataFrame
        for column_name in city_records.columns:
            if isinstance(contact_column, list):  # Check if the contact_column is a list (indicating multiple columns)
                if column_name.startswith(contact_column[0] + '_'):  # Check if the column name starts with the contact_column prefix
                    duplicate_column_names.append(column_name)
            elif column_name.startswith(contact_column + '_'):  # Check if the column name starts with the contact_column prefix
                duplicate_column_names.append(column_name)

        updated_contact_city_columns.append(duplicate_column_names)  # Append the list of duplicate column names to updated_contact_city_columns
    duplicated_col_max_len = max([len(t) for t in updated_contact_city_columns])  # Find the maximum length of the duplicate column names
    column_data = [col_type[0] for col_type in column_type]  # Extract the column types from the column_type list

    contact_matched_columns_list=[]
    contact_matched_columns_type=[]
    names_column=[]
    city_records.replace('', np.nan, inplace=True)
    for j in range(1,duplicated_col_max_len+1):
        for i in range(len(contact_city_columns)):
            if type(contact_city_columns[i])==list:
                name_dup_list=[]
                contact_matched_columns_list.append(contact_city_columns[i][0]+'_'+str(j))
                name_dup_list.append(contact_city_columns[i][0]+'_'+str(j))
                if contact_city_columns[i][1]!='':
                    name_dup_list.append(contact_city_columns[i][1]+'_'+str(j))
                else:
                    name_dup_list.append('')
                if contact_city_columns[i][2]!='':
                    name_dup_list.append(contact_city_columns[i][2]+'_'+str(j))
                else:
                    name_dup_list.append('')
                names_column.append(name_dup_list)
            else:
                contact_matched_columns_list.append(contact_city_columns[i]+'_'+str(j))
            contact_matched_columns_type.append(column_type[i]) 
    print(names_column)
    contact_matched_records = city_records.loc[:, contact_matched_columns_list]
    contact_matched_records_columns_type= [col_type[0] for col_type in contact_matched_columns_type]
    name_email_columns=[]
    name_phone_columns=[]
    name_index=[]
    for i in range(len(contact_matched_records_columns_type)):
        if contact_matched_records_columns_type[i]=='Name':
            name_index.append(i)
            for j in range(i+1,len(contact_matched_records_columns_type)):
                if contact_matched_records_columns_type[j]=='Email':
                    email=[contact_matched_columns_list[i],contact_matched_columns_list[j]]
                    name_email_columns.append(email)
                elif contact_matched_records_columns_type[j]=='Phone':   
                    phone=[contact_matched_columns_list[i],contact_matched_columns_list[j]]
                    name_phone_columns.append(phone)
                elif contact_matched_records_columns_type[j]=='Name':
                    break              
    contact_matched_records.fillna('',inplace=True)
    for x in range(len(name_email_columns)-1):
        df=pd.DataFrame(contact_matched_records[name_email_columns[x]])
        df.fillna('',inplace=True)                                
        for y in range(x+1,len(name_email_columns)):
            df1=pd.DataFrame(contact_matched_records[name_email_columns[y]])
            df1.fillna('',inplace=True)
            df1.astype(str)
            for j in range(df1.shape[0]):
                if (str(string_filter_for_abbreviation(df[name_email_columns[x][0]][j])) == str(string_filter_for_abbreviation(df1[name_email_columns[y][0]][j]))) or df[name_email_columns[x][0]][j] == '' or df1[name_email_columns[y][0]][j] == '':
                    if str(string_filter_for_abbreviation(df[name_email_columns[x][1]][j]))==str(string_filter_for_abbreviation(df1[name_email_columns[y][1]][j])):
                        contact_matched_records[name_email_columns[y][1]][j]=''

    contact_matched_records.fillna('',inplace=True)        
    email_columns=[name_email[1] for name_email in name_email_columns]
    for v in email_columns:
        contact_matched_records.loc[~contact_matched_records[v].str.contains('@'), v] = ''
        contact_matched_records[v] = contact_matched_records[v].replace('', np.nan)                    
    
    for x in range(len(name_phone_columns)-1):
        df=pd.DataFrame(contact_matched_records[name_phone_columns[x]])
        df.fillna('',inplace=True)                              
        for y in range(x+1,len(name_phone_columns)):
            df1=pd.DataFrame(contact_matched_records[name_phone_columns[y]])
            df1.fillna('',inplace=True)
            df1.astype(str)
            for j in range(df1.shape[0]):
                if str(phone_number_formatting(df[name_phone_columns[x][0]][j]))==str(phone_number_formatting(df1[name_phone_columns[y][0]][j])) or str(phone_number_formatting(df[name_phone_columns[x][0]][j]))=='' or str(phone_number_formatting(df1[name_phone_columns[y][0]][j]))=='':
                    if str(phone_number_formatting(df[name_phone_columns[x][1]][j]))==str(phone_number_formatting(df1[name_phone_columns[y][1]][j])):
                        contact_matched_records[name_phone_columns[y][1]][j]=''

    contact_matched_records.fillna('',inplace=True)
    index_num=0
    if matched_records=='business_matched':
        contact_id_column_index = city_records.columns.get_loc('ID')
        contact_city_column_index = city_records.columns.get_loc('city_index')
        if contact_city_column_index-contact_id_column_index>1:
            bludot_contact_records_indexes=[i for i in range(contact_id_column_index,contact_city_column_index)]
            bludot_contact_records=city_records.iloc[:,bludot_contact_records_indexes]            
            bludot_contact_records.fillna('',inplace=True)
            bludot_contact_columns=bludot_contact_records.columns
            l1=[]
            for i in range(len(bludot_contact_columns)):
                if bludot_contact_columns[i]=='Name' or 'Name.' in bludot_contact_columns[i] :
                    for j in range(i+1,len(bludot_contact_columns)):
                        y=bludot_contact_columns[j]
                        if 'Name.' in y:
                            break
                        elif y=='Contact' or 'Contact.' in bludot_contact_columns[j]:
                            l2=[]
                            l2.append(bludot_contact_columns[i])
                            l2.append(y)
                            l1.append(l2)
                            
            for j1 in l1:
                for x in range(len(name_email_columns)-1):
                    df=pd.DataFrame(contact_matched_records[name_email_columns[x]])
                    df.fillna('',inplace=True)
                    for j in range(df.shape[0]):
                        if (str(string_filter_for_abbreviation(df[name_email_columns[x][0]][j])) == str(string_filter_for_abbreviation(bludot_contact_records[j1[0]][j]))) or df[name_email_columns[x][0]][j] == '' or bludot_contact_records[j1[0]][j] == '':
                            if str(string_filter_for_abbreviation(df[name_email_columns[x][1]][j]))==str(string_filter_for_abbreviation(bludot_contact_records[j1[1]][j])):
                                if bludot_contact_records[j1[0]][j]!='':
                                    contact_matched_records[name_email_columns[x][1]][j]=''
                                
                    contact_matched_records.fillna('',inplace=True)        
                    email_columns=[name_email[1] for name_email in name_email_columns]
                    for v in email_columns:
                        contact_matched_records.loc[~contact_matched_records[v].str.contains('@'), v] = ''
                        contact_matched_records[v] = contact_matched_records[v].replace('', np.nan)                    

                for x in range(len(name_phone_columns)-1):
                    df=pd.DataFrame(contact_matched_records[name_phone_columns[x]])
                    df.fillna('',inplace=True)
                    for j in range(df.shape[0]):
                        if str(phone_number_formatting(df[name_phone_columns[x][0]][j]))==str(phone_number_formatting(bludot_contact_records[j1[0]][j])) or str(phone_number_formatting(df[name_phone_columns[x][0]][j]))=='' or str(phone_number_formatting(bludot_contact_records[j1[0]][j]))=='':
                            if str(phone_number_formatting(df[name_phone_columns[x][1]][j]))==str(phone_number_formatting(bludot_contact_records[j1[1]][j])):
                                if bludot_contact_records[j1[0]][j]!='':
                                    contact_matched_records[name_phone_columns[x][1]][j]=''

                            
    contact_matched_records.fillna('',inplace=True)         
    while index_num<contact_matched_records.shape[1]:
        if contact_matched_columns_type[index_num][0]=='Name':
            index_num_name = contact_matched_columns_list.index(contact_matched_columns_list[index_num])
            for name_list in names_column:
                if contact_matched_columns_list[index_num]==name_list[0]:
                    if name_list[1]!='':
                        contact_matched_records.insert(index_num_name+1,f'Title_{index_num_name+1}',city_records[name_list[1]])
                        contact_matched_columns_type.insert(index_num_name+1,['Title'])
                        contact_matched_columns_list.insert(index_num_name+1,name_list[1])
                    else:
                        contact_matched_records.insert(index_num_name+1,f'Title_{index_num_name+1}',contact_matched_columns_type[index_num][1])
                        contact_matched_columns_type.insert(index_num_name+1,['Title'])
                        contact_matched_columns_list.insert(index_num_name+1,f'Title_{index_num_name}')
                    
                    if name_list[2]!='':
                        contact_matched_records.insert(index_num_name+2,f'Roles_{index_num_name+2}',city_records[name_list[2]])
                        contact_matched_columns_type.insert(index_num_name+2,['Roles'])
                        contact_matched_columns_list.insert(index_num_name+2,name_list[2])
                    else:
                        contact_matched_records.insert(index_num_name+2,f'Roles_{index_num_name+2}',contact_matched_columns_type[index_num][2])
                        contact_matched_columns_type.insert(index_num_name+2,['Roles'])
                        contact_matched_columns_list.insert(index_num_name+2,f'Roles_{index_num_name}')
     
            for i, column in enumerate(contact_matched_columns_list):
                if column in city_records.columns:
                     contact_matched_records.loc[(contact_matched_records[contact_matched_columns_list[index_num]] == ''), [f'Title_{index_num_name+1}',f'Roles_{index_num_name+2}']] = ''
            index_num=index_num+3
        
        elif contact_matched_columns_type[index_num][0]=='Email' or contact_matched_columns_type[index_num][0]=='Phone':
            index_num_email = contact_matched_columns_list.index(contact_matched_columns_list[index_num])
            contact_matched_records.insert(index_num_email+1,f'Contact_type_{index_num_email+1}',contact_matched_columns_type[index_num][1])
            contact_matched_columns_type.insert(index_num_email+1,['Contact_type'])
            contact_matched_columns_list.insert(index_num_email+1,f'Contact_type_{index_num_email}')
            contact_matched_records.insert(index_num_email+2,f'Type_{index_num_email+2}',contact_matched_columns_type[index_num][2])
            contact_matched_columns_type.insert(index_num_email+2,['Type'])
            contact_matched_columns_list.insert(index_num_email+1,f'Contact_type_{index_num_email}')
                
            for i, column in enumerate(contact_matched_columns_list):
                if column in city_records.columns:
                    contact_matched_records.loc[(contact_matched_records[contact_matched_columns_list[index_num]] == ''), [f'Contact_type_{index_num_email+1}',f'Type_{index_num_email+2}']] = ''
            index_num=index_num+3
        else:
            index_num=index_num+1
    column_name_contact_sheet=[col_type[0] for col_type in contact_matched_columns_type]
    column_name_contact_sheet = ['Contact' if col_name == 'Email' or col_name=='Phone' else col_name for col_name in column_name_contact_sheet]
    indices = [index for index, sublist in enumerate(contact_matched_columns_type) if 'Name' in sublist]
    for i in range(len(indices)-1):
        for j in range(i+1, len(indices)):
            for x in range(contact_matched_records.shape[0]):
                if str(string_filter_for_abbreviation(contact_matched_records.iloc[x, indices[i]])) == str(string_filter_for_abbreviation(contact_matched_records.iloc[x, indices[j]])):
                    if contact_matched_records.iloc[x, indices[i]] != "":
                        if indices[j] != indices[-1]:
                            columns_to_convert=contact_matched_records.columns[indices[j]:indices[j+1]]
                            row_list = contact_matched_records.loc[x, columns_to_convert].tolist()
                            flag=True
                            for y in row_list[1:]:
                                if y!='':
                                    flag=False
                                    break
                                
                            if flag:
                                contact_matched_records.iloc[x, indices[j]]=''
                            
                            else:
                                columns_to_convert=contact_matched_records.columns[indices[i]:indices[i+1]]
                                row_list = contact_matched_records.loc[x, columns_to_convert].tolist()
                                flag=True
                                for y in row_list[1:]:
                                    if y!='':
                                        flag=False
                                        break
                                if flag:
                                    contact_matched_records.iloc[x, indices[i]]=''

                                    

    for x in range(len(indices)-1):
        df2 = contact_matched_records.loc[:, [contact_matched_records.columns[i] for i in range(indices[x], indices[x+1])]]

        # Assuming your DataFrame is named df
        column_to_check = df2.columns[0]

        df2[column_to_check] = df2[column_to_check].fillna('')  # Replace NaN values with empty string

        # Get all columns except the first column
        other_columns = df2.columns[1:]

        # Check conditions and update the first column
        conditions = (df2[column_to_check] == '') & (df2[other_columns] != '').any(axis=1)
        df2.loc[conditions, column_to_check] = '-'
        contact_matched_records[df2.columns[0]]=df2[df2.columns[0]]
    contact_matched_records.columns=column_name_contact_sheet
    empty_cols_to_drop=[]
    for col_index,col_name in  enumerate(contact_matched_records.columns):
        contact_matched_records.iloc[:,[col_index]] = contact_matched_records.iloc[:,[col_index]].replace('', None)
        col_is_null = np.all(pd.isnull(contact_matched_records.iloc[:,[col_index]]))
        if col_name=='Name':
            if col_is_null:
                empty_cols_to_drop.extend([col_index,col_index+1,col_index+2])
        elif col_name=='Title' or col_name=='Roles':
            continue
        else:
            if col_is_null:
                empty_cols_to_drop.append(col_index)
    for index in empty_cols_to_drop:
        column_name_contact_sheet[index]='drop_col'
    contact_matched_records.columns=column_name_contact_sheet
    columns_to_drop = [col for col in contact_matched_records.columns if col == 'drop_col']
    contact_matched_records = contact_matched_records.drop(columns=columns_to_drop)
    
    if matched_records=='additional_matched':
        contact_matched_records.to_excel(fr"C:\Users\Kavita Patel\Desktop\new_city_onboarding_project\new_city_onboarding\cities_and_counties\{CITY_NAME}\Results\Output\Final Excel\Additional_Contact_Matched_Records.xlsx",index=False)
    elif matched_records=='business_matched':
        if contact_city_column_index-contact_id_column_index>1:
            updated_bludot_columns_name=[]
            for i in bludot_contact_records.columns:
                if '.' in i:
                    k=i.split('.')
                    updated_bludot_columns_name.append(k[0])
                else:
                    updated_bludot_columns_name.append(i)
            bludot_contact_records.columns=updated_bludot_columns_name
            empty_cols_to_drop=[]
            updated_col_with_drop_col=list(bludot_contact_records.columns)
            for col_index,col_name in  enumerate(bludot_contact_records.columns):
                bludot_contact_records.iloc[:,[col_index]] = bludot_contact_records.iloc[:,[col_index]].replace('', None)
                col_is_null = np.all(pd.isnull(bludot_contact_records.iloc[:,[col_index]]))
                if col_name=='Name':
                    if col_is_null:
                        empty_cols_to_drop.extend([col_index,col_index+1,col_index+2])
                elif col_name=='Title' or col_name=='Roles':
                    continue
                else:
                    if col_is_null:
                        empty_cols_to_drop.append(col_index)
            for index in empty_cols_to_drop:
                updated_col_with_drop_col[index]='drop_col'
            bludot_contact_records.columns=updated_col_with_drop_col
            columns_to_drop = [col for col in bludot_contact_records.columns if col == 'drop_col']
            bludot_contact_records = bludot_contact_records.drop(columns=columns_to_drop)
            bludot_contact_records.fillna('',inplace=True)
            contact_matched_records.fillna('',inplace=True)
            bludot_indices=[index for index,col in enumerate(list(bludot_contact_records.columns)) if col == 'Name']
            city_indices=[index for index, sublist in enumerate(list(contact_matched_records.columns)) if 'Name' in sublist]
            for i in range(len(bludot_indices)):
                for j in range(len(city_indices)):
                    for x in range(contact_matched_records.shape[0]):
                        if str(string_filter_for_abbreviation(bludot_contact_records.iloc[x, bludot_indices[i]])) == str(string_filter_for_abbreviation(contact_matched_records.iloc[x, city_indices[j]])):
                            # print(str(string_filter_for_abbreviation(bludot_contact_records.iloc[x, bludot_indices[i]])),str(string_filter_for_abbreviation(contact_matched_records.iloc[x, city_indices[j]])))
                            if bludot_contact_records.iloc[x, bludot_indices[i]] != "":
                                if city_indices[j] != city_indices[-1]:
                                    row_list = contact_matched_records.iloc[x, city_indices[j]:city_indices[j]+1].tolist()
                                    flag=True
                                    for y in row_list[1:]:
                                        if y!='':
                                            flag=False
                                            break
                                    if flag:
                                        contact_matched_records.iloc[x, city_indices[j]]=''
                
            contact_matched_records=pd.concat([bludot_contact_records, contact_matched_records], axis=1)
            contact_matched_records.fillna('',inplace=True)
            contact_matched_records.to_excel(fr"C:\Users\Kavita Patel\Desktop\new_city_onboarding_project\new_city_onboarding\cities_and_counties\{CITY_NAME}\results\output\final_excel\Contact_Matched_Records.xlsx",index=False)                
            
        else:
            contact_matched_records.to_excel(fr"C:\Users\Kavita Patel\Desktop\new_city_onboarding_project\new_city_onboarding\cities_and_counties\{CITY_NAME}\results\output\final_excel\Contact_Matched_Records.xlsx",index=False)